"""
生成句子
"""

import re
import os
import copy

import openpyxl

import app.mt.commonApi as commonApi
import app.mt.common as common

from app.mt.syntactic import find_clauses
from app.mt.cc import cc_part
from app.mt.prune import pruning


def gen(sentence):
    """
    单个句子剪枝结果生成
    :param sentence: 语句
    :return: 扩增结果
    """
    sentence = correct_punctuation(sentence)
    token = common.nlpEN.word_tokenize(sentence)
    # 切分句子为主句和从句
    sent_main, sent_clauses = find_clauses(sentence)

    # 存在主句
    if len(sent_main) != 0:
        if sent_main[0] == "," or sent_main[0] == ".":
            sent_main = sent_main[1 : len(sent_main)]
        # 主句直接添加为生成句
        if len(sent_clauses) != 0:
            common.sentences_gen.append(sent_main)
        token_main = common.nlpEN.word_tokenize(sent_main)
        dependency_main = common.nlpEN.dependency_parse(sent_main)
        # dependency_main_enhancep = common.nlpEN.dependency_parse_enhancep(sent_main)
        # 寻找主句中需要保留的部分
        common.retain_tokens.clear()
        common.retain_tokens = commonApi.find_inseparable(token_main)
        # 识别主干内容
        common.trunk.clear()
        commonApi.dependency_traversal(dependency_main, token_main, common.trunk)
        # 识别主句中的并列成分
        cc_list = cc_part(
            dependency_main,
            commonApi.IndexTree.fromstring(common.nlpEN.parse(sent_main)),
        )
        # 主句无并列
        if len(cc_list) == 0:
            # 构建依存关系树
            dependency_tree_main, root = commonApi.construct_dependency_tree(
                dependency_main, token_main
            )
            remain_token_main = [i for i in range(len(token_main))]
            pruning(
                dependency_tree_main,
                root,
                dependency_main,
                token_main,
                remain_token_main,
                sent_main,
            )
        # 主句有并列成分
        else:
            for sent_cc in cc_list:
                ch = sent_cc[-1]
                if ch not in common.punctuation_en:
                    sent_cc = sent_cc + "."
                common.sentences_gen.append(sent_cc)
                token_cc = common.nlpEN.word_tokenize(sent_cc)
                dependency_cc = common.nlpEN.dependency_parse(sent_cc)
                # 寻找并列句中需要保留的部分
                common.retain_tokens.clear()
                common.retain_tokens = commonApi.find_inseparable(token_cc)
                # 识别并列句中的主干内容
                common.trunk.clear()
                commonApi.dependency_traversal(dependency_cc, token_cc, common.trunk)
                # 构建依存关系树
                dependency_tree_cc, root = commonApi.construct_dependency_tree(
                    dependency_cc, token_cc
                )
                remain_token_cc = [i for i in range(len(token_cc))]
                pruning(
                    dependency_tree_cc,
                    root,
                    dependency_cc,
                    token_cc,
                    remain_token_cc,
                    sent_cc,
                )

    # 对从句进行处理，步骤与主句相同
    for sent_clause in sent_clauses:
        common.sentences_gen.append(sent_clause)
        common.trunk.clear()
        clause_dependency = common.nlpEN.dependency_parse(sent_clause)
        # clause_dependency_enhancep = common.nlpEN.dependency_parse_enhancep(sent_clause)
        clause_token = common.nlpEN.word_tokenize(sent_clause)
        common.retain_tokens = commonApi.find_inseparable(clause_token)
        commonApi.dependency_traversal(clause_dependency, clause_token, common.trunk)
        cc_list = cc_part(
            clause_dependency,
            commonApi.IndexTree.fromstring(common.nlpEN.parse(sent_clause)),
        )

        if len(cc_list) == 0:
            clause_dependency_tree, root = commonApi.construct_dependency_tree(
                clause_dependency, clause_token
            )
            remain_token_clause = [i for i in range(len(clause_token))]
            pruning(
                clause_dependency_tree,
                root,
                clause_dependency,
                clause_token,
                remain_token_clause,
                sent_clause,
            )
        else:
            for sent_cc in cc_list:
                ch = sent_cc[-1]
                if ch not in common.punctuation_en:
                    sent_cc = sent_cc + "."
                common.sentences_gen.append(sent_cc)
                token_cc = common.nlpEN.word_tokenize(sent_cc)
                dependency_cc = common.nlpEN.dependency_parse(sent_cc)
                common.retain_tokens.clear()
                common.retain_tokens = commonApi.find_inseparable(token_cc)
                common.trunk.clear()
                commonApi.dependency_traversal(dependency_cc, token_cc, common.trunk)
                dependency_tree_cc, root = commonApi.construct_dependency_tree(
                    dependency_cc, token_cc
                )
                remain_token_cc = [i for i in range(len(token_cc))]
                pruning(
                    dependency_tree_cc,
                    root,
                    dependency_cc,
                    token_cc,
                    remain_token_cc,
                    sent_cc,
                )
    result = copy.deepcopy(common.sentences_gen)

    if len(token) > 6 and len(result) == 0:
        if sentence.find(":") != -1:
            sent11 = sentence.split(":")
            l1 = gen(sent11[0] + ".")
            l2 = gen(sent11[1])
            l1.append(sent11[1])
            for i in l2:
                l1.append(i)
            return l1
    result2 = [correct_punctuation(r.strip()) for r in result]
    result3 = []
    [result3.append(i) for i in result2 if i not in result3]
    return result3


def correct_punctuation(sentence: str):
    """
    修正标点符号
    :param sentence:
    :return:
    """
    punc_pattern = re.compile("[,.]{2,}")
    s = sentence
    # 加句号
    if not s.endswith('."') and not s.endswith("."):
        s += "."

    # 去除多个标点
    result = re.search(punc_pattern, s)
    while result is not None:
        span = result.span()
        # 在句首直接去除
        if span[0] == 0:
            s = s[span[1] :]
        # 在句末保留最后一个
        elif span[1] == len(s):
            s = s[: span[0]] + s[-1]
        # 在句中保留第一个
        else:
            s = s[: span[0] + 1] + s[span[1] :]
        result = re.search(punc_pattern, s)

    # 删除未能匹配成对的双引号
    last_index_quote_mark = s.find('"', 0)
    quote_mark = 1
    while last_index_quote_mark >= 0:
        tmp_last = last_index_quote_mark
        last_index_quote_mark = s.find('"', tmp_last + 1)
        if last_index_quote_mark < 0:
            if quote_mark % 2 == 1:
                s = s[:tmp_last] + s[tmp_last + 1 :]
            break
        else:
            quote_mark += 1

    # 去除开头的标点
    while len(s) > 0 and (s[0] == "." or s[0] == ","):
        s = s[1:]

    return s.strip()


# 生成所有数据
def gen_all(dataset):
    """
    数据集数据生成
    :param dataset: 数据集path
    :return:
    """
    commonApi.read_conf()
    nlp_en = common.nlpEN
    dic = {}
    str_list = []

    workbook = openpyxl.load_workbook("./data/News.xlsx")
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "./gcd.json"  # 凭据
    for name in workbook.sheetnames:
        if name != dataset:
            continue
        sheet = workbook[name]  # 获取指定sheet表
        sheet_len = eval(sheet.dimensions.split("B")[1])
        for i in range(sheet_len):
            cell = str(sheet.cell(row=i + 1, column=1).value)
            line = cell
            sent = line.split("\n")[0]
            dependency = nlp_en.dependency_parse(sent)
            token = nlp_en.word_tokenize(sent)
            strr = ""
            for i, begin, end in dependency:
                if begin - 1 < 0:
                    first = "NULL"
                else:
                    first = token[begin - 1]
                last = token[end - 1]
                strr += (
                    i
                    + "-".join([str(begin), first])
                    + "-".join([str(end), last])
                    + "\n"
                )
            str_list.append(strr)
            phrases = gen(sent)
            dic[sent] = []
            for t in phrases:
                dic[sent].append(t)
    return dic, str_list
