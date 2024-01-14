"""
Handles generating sentences.
"""

import copy
import re

import commonApi
import common

from syntactic import find_clauses
from cc import cc_part
from prune import pruning
import os
import openpyxl


def gen(sentence):
    """
    Takes a sentence as input, performs sentence pruning and returns the result.

    Tokenizes the input sentence, identifies the main and subordinate clauses, and
    prunes unnecessary parts. It handles cases where the sentence includes a colon by
    splitting it into two parts and recursively processing them.
    The corrected sentences are stored in the common.sentences_gen list.
    """
    sentence = correct_punctuation(sentence)
    token = common.nlpEN.word_tokenize(sentence)
    # Split sentences into main clauses and subordinate clauses.
    sent_main, sent_clauses = find_clauses(sentence)

    if len(sent_main) != 0:  # If main clause exists
        if sent_main[0] == "," or sent_main[0] == ".":
            sent_main = sent_main[1 : len(sent_main)]

        if len(sent_clauses) != 0:
            # The main clause is directly added as a generated sentence.
            common.sentences_gen.append(sent_main)
        token_main = common.nlpEN.word_tokenize(sent_main)
        dependency_main = common.nlpEN.dependency_parse(sent_main)
        # dependency_main_enhancep = common.nlpEN.dependency_parse_enhancep(sent_main)
        # Find the part of the main clause that needs to be retained.
        common.retain_tokens.clear()
        common.retain_tokens = commonApi.find_inseparable(token_main)
        # Identify the main content.
        common.trunk.clear()
        commonApi.dependency_traversal(dependency_main, token_main, common.trunk)

        cc_list = cc_part(
            dependency_main,
            commonApi.IndexTree.fromstring(common.nlpEN.parse(sent_main)),
        )
        # 主句无并列
        if len(cc_list) == 0:
            # Build dependency tree.
            dependency_tree_main, root = commonApi.construct_dependency_tree(
                dependency_main, token_main
            )
            remain_token_main = [i for i in range(len(token_main))]
            pruning(
                dependency_tree_main,
                root,
                dependency_main,
                token_main,
                remain_token_main,
                sent_main,
            )
        # 主句有并列成分
        else:
            for sent_cc in cc_list:
                ch = sent_cc[-1]
                if ch not in common.punctuation_en:
                    sent_cc = sent_cc + "."
                common.sentences_gen.append(sent_cc)
                token_cc = common.nlpEN.word_tokenize(sent_cc)
                dependency_cc = common.nlpEN.dependency_parse(sent_cc)
                # 寻找并列句中需要保留的部分
                common.retain_tokens.clear()
                common.retain_tokens = commonApi.find_inseparable(token_cc)
                # 识别并列句中的主干内容
                common.trunk.clear()
                commonApi.dependency_traversal(dependency_cc, token_cc, common.trunk)
                # Build dependency tree.
                dependency_tree_cc, root = commonApi.construct_dependency_tree(
                    dependency_cc, token_cc
                )
                remain_token_cc = [i for i in range(len(token_cc))]
                pruning(
                    dependency_tree_cc,
                    root,
                    dependency_cc,
                    token_cc,
                    remain_token_cc,
                    sent_cc,
                )

    # Process the subordinate clause, the steps are the same as for the main clause.
    for sent_clause in sent_clauses:
        common.sentences_gen.append(sent_clause)
        common.trunk.clear()
        clause_dependency = common.nlpEN.dependency_parse(sent_clause)
        # clause_dependency_enhancep = common.nlpEN.dependency_parse_enhancep(sent_clause)
        clause_token = common.nlpEN.word_tokenize(sent_clause)
        common.retain_tokens = commonApi.find_inseparable(clause_token)
        commonApi.dependency_traversal(clause_dependency, clause_token, common.trunk)
        cc_list = cc_part(
            clause_dependency,
            commonApi.IndexTree.fromstring(common.nlpEN.parse(sent_clause)),
        )

        if len(cc_list) == 0:
            clause_dependency_tree, root = commonApi.construct_dependency_tree(
                clause_dependency, clause_token
            )
            remain_token_clause = [i for i in range(len(clause_token))]
            pruning(
                clause_dependency_tree,
                root,
                clause_dependency,
                clause_token,
                remain_token_clause,
                sent_clause,
            )
        else:
            for sent_cc in cc_list:
                ch = sent_cc[-1]
                if ch not in common.punctuation_en:
                    sent_cc = sent_cc + "."
                common.sentences_gen.append(sent_cc)
                token_cc = common.nlpEN.word_tokenize(sent_cc)
                dependency_cc = common.nlpEN.dependency_parse(sent_cc)
                common.retain_tokens.clear()
                common.retain_tokens = commonApi.find_inseparable(token_cc)
                common.trunk.clear()
                commonApi.dependency_traversal(dependency_cc, token_cc, common.trunk)
                dependency_tree_cc, root = commonApi.construct_dependency_tree(
                    dependency_cc, token_cc
                )
                remain_token_cc = [i for i in range(len(token_cc))]
                pruning(
                    dependency_tree_cc,
                    root,
                    dependency_cc,
                    token_cc,
                    remain_token_cc,
                    sent_cc,
                )
    result = copy.deepcopy(common.sentences_gen)

    if len(token) > 6 and len(result) == 0:
        if sentence.find(":") != -1:
            sent11 = sentence.split(":")
            l1 = gen(sent11[0] + ".")
            l2 = gen(sent11[1])
            l1.append(sent11[1])
            for i in l2:
                l1.append(i)
            return l1
    result2 = [correct_punctuation(r.strip()) for r in result]
    result3 = []
    [result3.append(i) for i in result2 if i not in result3]
    return result3


def correct_punctuation(sentence: str) -> str:
    """
    Corrects punctuation in given sentence and returns the obtained sentence.
    """
    punc_pattern = re.compile("[,.]{2,}")
    s = sentence
    # Add period.
    if not s.endswith('."') and not s.endswith("."):
        s += "."

    # Remove multiple punctuation marks.
    result = re.search(punc_pattern, s)
    while result is not None:
        span = result.span()
        if span[0] == 0:
            # Remove the symbol from the beginning of the sentence.
            s = s[span[1] :]
        elif span[1] == len(s):
            # Only keep the last appearance, at the end of the sentence.
            s = s[: span[0]] + s[-1]
        else:
            # Keep the first appearance in the sentence.
            s = s[: span[0] + 1] + s[span[1] :]
        result = re.search(punc_pattern, s)

    # Remove unmatched double quotes.
    last_index_quote_mark = s.find('"', 0)
    quote_mark = 1
    while last_index_quote_mark >= 0:
        tmp_last = last_index_quote_mark
        last_index_quote_mark = s.find('"', tmp_last + 1)
        if last_index_quote_mark < 0:
            if quote_mark % 2 == 1:
                s = s[:tmp_last] + s[tmp_last + 1 :]
            break
        else:
            quote_mark += 1

    # Remove leading punctuation.
    while len(s) > 0 and (s[0] == "." or s[0] == ","):
        s = s[1:]

    return s.strip()


def gen_all(dataset_path: str):
    """
    Generates pruned sentences for a given dataset (all data).
    """
    commonApi.read_conf()
    nlp_en = common.nlpEN
    dic = {}
    str_list = []

    workbook = openpyxl.load_workbook("./data/News.xlsx")
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "./gcd.json"  # 凭据
    for name in workbook.sheetnames:
        if name != dataset_path:
            continue
        sheet = workbook[name]  # 获取指定sheet表
        sheet_len = eval(sheet.dimensions.split("B")[1])
        for i in range(sheet_len):
            cell = str(sheet.cell(row=i + 1, column=1).value)
            line = cell
            sent = line.split("\n")[0]
            dependency = nlp_en.dependency_parse(sent)
            token = nlp_en.word_tokenize(sent)
            strr = ""
            for i, begin, end in dependency:
                if begin - 1 < 0:
                    first = "NULL"
                else:
                    first = token[begin - 1]
                last = token[end - 1]
                strr += (
                    i
                    + "-".join([str(begin), first])
                    + "-".join([str(end), last])
                    + "\n"
                )
            str_list.append(strr)
            phrases = gen(sent)
            dic[sent] = []
            for t in phrases:
                dic[sent].append(t)
    return dic, str_list


if __name__ == "__main__":
    commonApi.read_conf()
    # sent = "he said he expected that fiscal policy will be supportive and the macro leverage ratio will rise modestly next year."
    sent = 'Ni Pengfei, a research center director of the Chinese Academy of Social Sciences, was quoted by China News Service as saying that "policies must be stable and consistent, and at the same time must be fine-tuned according to different areas\' specific situations".'
    print("\n".join(gen(sent)))
    common.nlpEN.close()
