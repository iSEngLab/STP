import xlrd
import string
from stanfordcorenlp import StanfordCoreNLP
import openpyxl
from gensim import corpora

ssss = [
    "Politics",
    "Business",
    "Culture",
    "Sports",
    "Tech",
    "TravelFood",
    "Health",
    "LifeStyle",
    "Legal",
    "Opinion",
]
software = ["Google", "Bing"]
workbook = openpyxl.load_workbook("./data/News.xlsx")
origin_source_sentsL = []

for name in workbook.sheetnames:
    sheet = workbook[name]  # 获取指定sheet表
    sheet_len = eval(sheet.dimensions.split("B")[1])
    for i in range(sheet_len):
        cell = str(sheet.cell(row=i + 1, column=1).value)
        line = cell
        sent = line.split("\n")[0]
        origin_source_sentsL.append(sent)


STPBingList = []
STPGoogleList = []
STPDeepLList = []
SITBingList = []
SITGoogleList = []
PatInvBingList = []
PatInvGoogleList = []
RTIGoogleList = []
RTIBingList = []


nlpCN = StanfordCoreNLP(r"./stanford-corenlp-4.5.5", lang="zh")
nlpEN = StanfordCoreNLP(r"./stanford-corenlp-4.5.5")


class Pair:
    def __init__(
        self,
        source,
        gen,
        transl,
        transgl,
        error,
        serror,
        gerror,
        distance,
        distance_word,
    ):
        self.source = source
        self.gen = gen
        self.transl = transl
        self.trangl = transgl
        self.error = error
        self.serror = serror
        self.gerror = gerror
        self.distance = distance
        self.distance_word = distance_word


class SentT:
    def __init__(self, sent, trans, error):
        self.sent = sent
        self.trans = trans
        self.error = error


word_token_en = {}
word_token_cn = {}


def JustSub(ll1, ll2):
    map1 = {}
    l1 = []
    if ll1 not in word_token_en.keys():
        l1 = nlpEN.word_tokenize(ll1)
        word_token_en[ll1] = l1
    else:
        l1 = word_token_en[ll1]
    l2 = []
    if ll2 not in word_token_en.keys():
        l2 = nlpEN.word_tokenize(ll2)
        word_token_en[ll2] = l2
    else:
        l2 = word_token_en[ll2]
    punc = string.punctuation
    word_sent1_count = 0
    for i in l1:
        if i in punc:  # type: ignore
            continue
        word_sent1_count = word_sent1_count + 1
        count = map1.get(i, 0)
        map1[i] = count + 1

    map2 = {}
    word_sent2_count = 0
    for i in l2:
        if i in punc:  # type: ignore
            continue
        word_sent2_count = word_sent2_count + 1
        count = map2.get(i, 0)
        map2[i] = count + 1
        if map1.get(i, -1) == 0 or map1.get(i, 0) < map2.get(i, 0):
            return False, abs(word_sent2_count - word_sent1_count)
    return True, abs(word_sent2_count - word_sent1_count)


def out():
    # software=["google","bing"]
    # ssss=['politics', 'business']
    #
    ssss = [
        "Politics",
        "Business",
        "Culture",
        "Sports",
        "Tech",
        "TravelFood",
        "Health",
        "LifeStyle",
        "Legal",
        "Opinion",
    ]
    software = ["Google", "Bing"]
    result_all_dic = {}
    with open("stp.txt", "r", encoding="utf-8") as f:
        result_all_dic = eval(f.read())

    left_index = 0
    resultsent = []
    for item in software:
        for kkkk in ssss:
            res = ""
            for dis in range(left_index, 7):
                res = res + result_all_dic[kkkk][item][1][dis] + "\t"
            resultsent.append(res)

    for i in resultsent:
        print(i)


def bow(l1, l2):
    s1 = []
    if l1 not in word_token_en.keys():
        s1 = nlpEN.word_tokenize(l1)
        word_token_en[l1] = s1
    else:
        s1 = word_token_en[l1]
    s2 = []
    if l2 not in word_token_en.keys():
        s2 = nlpEN.word_tokenize(l2)
        word_token_en[l2] = s2
    else:
        s2 = word_token_en[l2]
    punc = string.punctuation
    text1 = []
    for i in s1:
        if i in punc:  # type: ignore
            continue
        text1.append(i)
    text2 = []
    for i in s2:
        if i in punc:  # type: ignore
            continue
        text2.append(i)
    texts = [text1, text2]
    dictionary = corpora.Dictionary(texts)
    # 利用doc2bow作为词袋模型
    corpus = [dictionary.doc2bow(text) for text in texts]
    dict1 = {}
    dict2 = {}
    for i, j in corpus[0]:
        dict1[i] = j
    for i, j in corpus[1]:
        dict2[i] = j
    dis = 0
    for i in range(len(dictionary.cfs.items())):
        if i not in dict1.keys():
            continue
        l1 = dict1.get(i, 0)
        l2 = dict2.get(i, 0)
        if l2 != l1:
            dis = dis + abs(l1 - l2)
    return dis


def bow1(l1, l2):
    s1 = []
    if l1 not in word_token_cn.keys():
        s1 = nlpCN.word_tokenize(l1)
        word_token_cn[l1] = s1
    else:
        s1 = word_token_cn[l1]
    s2 = []
    if l2 not in word_token_cn.keys():
        s2 = nlpCN.word_tokenize(l2)
        word_token_cn[l2] = s2
    else:
        s2 = word_token_cn[l2]
    punc = string.punctuation

    text1 = []
    for i in s1:
        if i in punc:  # type: ignore
            continue
        text1.append(i)
    text2 = []
    for i in s2:
        if i in punc:  # type: ignore
            continue
        text2.append(i)
    texts = [text1, text2]
    print(texts)
    dictionary = corpora.Dictionary(texts)
    # 利用doc2bow作为词袋模型
    corpus = [dictionary.doc2bow(text) for text in texts]
    dict1 = {}
    dict2 = {}
    for i, j in corpus[0]:
        dict1[i] = j
    for i, j in corpus[1]:
        dict2[i] = j
    dis = 0
    for i in text2:
        l1 = text1.count(i)
        l2 = text2.count(i)
        if l1 != l2:
            dis = dis + 1
    return dis


def PatInv(ErorShow=False):
    path = "./PatInv-Evaluated"
    print("PatInv")
    for item in software:
        print(item)
        for kkkk in ssss:
            paths = path + "/" + "PatInv_" + kkkk + "_" + item + ".xlsx"
            workbook = xlrd.open_workbook(paths)
            # 根据sheet索引或者名称获取sheet内容
            # print(paths)
            sheet1 = workbook.sheet_by_index(1)  # sheet索引从0开始

            Result = {}
            ListPair = []
            listYuan = []
            for i in range(1, sheet1.nrows - 1):
                type = i % 6
                if type == 1:
                    result1 = sheet1.row_values(i + 2)

                    result2 = sheet1.row_values(i + 5)

                    error1 = []
                    for kk in range(1, len(result1) - 1):
                        if len(str(result1[kk])) >= 1:
                            error1.append(kk)
                            if listYuan.count(result1[0]) < 1:
                                listYuan.append(result1[0])
                            if item.lower() == "google":
                                if PatInvGoogleList.count(result1[0]) < 1:
                                    PatInvGoogleList.append(result1[0])
                            else:
                                if PatInvBingList.count(result1[0]) < 1:
                                    PatInvBingList.append(result1[0])

                    error2 = []
                    Result[result1[0]] = SentT(result1[0], "", error1)
                    for kk in range(1, len(result2) - 1):
                        if len(str(result2[kk])) >= 1:
                            error2.append(kk)
                            if item.lower() == "google":
                                if PatInvGoogleList.count(result2[0]) < 1:
                                    PatInvGoogleList.append(result2[0])
                            else:
                                if PatInvBingList.count(result2[0]) < 1:
                                    PatInvBingList.append(result2[0])
                    Result[result2[0]] = SentT(result2[0], "", error2)
                    flag = False
                    if len(error1) > 0 or len(error2) > 0:
                        flag = True
                    ListPair.append(
                        Pair(
                            Result[result1[0]],
                            Result[result2[0]],
                            "",
                            "",
                            flag,
                            error1,
                            error2,
                            0,
                            0,
                        )
                    )

            ccc = 0
            i11 = 0
            for i in Result.keys():
                i11 += 1
                if len(Result[i].error) > 0:
                    ccc += 1
            # print(ll + "error:" + str(ccc) + " count:" + str(i11))
            ccc2 = 0
            i22 = 0
            for i in ListPair:
                i22 += 1
                if i.error is True:
                    ccc2 += 1

            # print(ll + "Pair error:" + str(ccc2) + " count:" + str(i22))
            # print("原句错误数量：" + str(len(listYuan)))
            if ErorShow is False:
                resTemp = "{:.1%}".format((ccc2 / i22))
                print(resTemp + "(%d/%d)" % (ccc2, i22))
            else:
                print("%d(%d)" % (ccc, len(listYuan)))


def RTI(ErorShow=False):
    path = "./RTI-Evaluated"
    print("RTI")
    for item in software:
        print(item)
        for kkkk in ssss:
            paths = path + "/" + "RTI_" + kkkk + "_" + item + ".xlsx"
            # print(paths)
            workbook = xlrd.open_workbook(paths)
            # 根据sheet索引或者名称获取sheet内容

            sheet1 = workbook.sheet_by_index(1)  # sheet索引从0开始

            # sheet1的名称，行数，列数
            ll = kkkk + "_" + item

            ll = ll[0]
            Result = {}
            listYuan = []
            ListPair = []

            for i in range(1, sheet1.nrows):
                type = i % 8
                if type == 1:
                    result1 = sheet1.row_values(i + 2)

                    result2 = sheet1.row_values(i + 5)

                    error1 = []
                    for kk in range(1, len(result1) - 1):
                        if len(str(result1[kk])) >= 1:
                            error1.append(kk)
                            if origin_source_sentsL.count(result1[0]) > 0:
                                if listYuan.count(result1[0]) < 1:
                                    listYuan.append(result1[0])
                            if item.lower() == "google":
                                if RTIGoogleList.count(result1[0]) < 1:
                                    RTIGoogleList.append(result1[0])
                            else:
                                if RTIBingList.count(result1[0]) < 1:
                                    RTIBingList.append(result1[0])

                    error2 = []
                    Result[result1[0]] = SentT(result1[0], "", error1)
                    for kk in range(1, len(result2) - 1):
                        if len(str(result2[kk])) >= 1:
                            error2.append(kk)
                            if item.lower() == "google":
                                if RTIGoogleList.count(result2[0]) < 1:
                                    RTIGoogleList.append(result2[0])
                            else:
                                if RTIBingList.count(result2[0]) < 1:
                                    RTIBingList.append(result2[0])
                    Result[result2[0]] = SentT(result2[0], "", error2)
                    flag = False
                    if len(error1) > 0 or len(error2) > 0:
                        flag = True
                    ListPair.append(
                        Pair(
                            Result[result1[0]],
                            Result[result2[0]],
                            "",
                            "",
                            flag,
                            error1,
                            error2,
                            0,
                            0,
                        )
                    )

            ccc = 0
            i11 = 0
            for i in Result.keys():
                i11 += 1
                if len(Result[i].error) > 0:
                    ccc += 1
            # print(ll + "error:" + str(ccc) + " count:" + str(i11))
            ccc2 = 0
            i22 = 0
            for i in ListPair:
                i22 += 1
                if i.error is True:
                    ccc2 += 1
            # print(ll + "Pair error:" + str(ccc2) + " count:" + str(i22))
            # print("原句错误数量：" + str(len(listYuan)))
            if ErorShow is False:
                resTemp = "{:.1%}".format((ccc2 / i22))
                print(resTemp + "(%d/%d)" % (ccc2, i22))
            else:
                print("%d(%d)" % (ccc, len(listYuan)))


def SIT(ErorShow=False):
    path = "./SIT-Evaluated"
    print("SIT:")
    for item in software:
        print(item)
        for kkkk in ssss:
            paths = path + "/" + "SIT_" + kkkk + "_" + item + ".xlsx"
            workbook = xlrd.open_workbook(paths)
            # 根据sheet索引或者名称获取sheet内容

            sheet1 = workbook.sheet_by_index(1)  # sheet索引从0开始

            Result = {}
            listYuan = []
            ListPair = []
            for i in range(1, sheet1.nrows):
                result = sheet1.row_values(i)  # 获取第三行

                if result[0].split(":")[0] == "Issue":
                    result1 = sheet1.row_values(i + 2)

                    result2 = sheet1.row_values(i + 6)

                    error1 = []
                    for kk in range(1, len(result1) - 1):
                        if len(str(result1[kk])) >= 1:
                            error1.append(kk)
                            if listYuan.count(result1[0]) < 1:
                                listYuan.append(result1[0])
                            if item.lower() == "google":
                                if SITGoogleList.count(result1[0]) < 1:
                                    STPGoogleList.append(result1[0])
                            else:
                                if SITBingList.count(result1[0]) < 1:
                                    STPBingList.append(result1[0])

                    error2 = []
                    Result[result1[0]] = SentT(result1[0], "", error1)
                    for kk in range(1, len(result2) - 1):
                        if len(str(result2[kk])) >= 1:
                            error2.append(kk)
                            if item.lower() == "google":
                                if STPGoogleList.count(result2[0]) < 1:
                                    STPGoogleList.append(result2[0])
                            else:
                                if STPBingList.count(result2[0]) < 1:
                                    STPBingList.append(result2[0])

                    Result[result2[0]] = SentT(result2[0], "", error2)

                    flag = False
                    if len(error1) > 0 or len(error2) > 0:
                        flag = True
                    ListPair.append(
                        Pair(
                            Result[result1[0]],
                            Result[result2[0]],
                            "",
                            "",
                            flag,
                            error1,
                            error2,
                            0,
                            0,
                        )
                    )

            ccc = 0
            i11 = 0
            for i in Result.keys():
                i11 += 1
                if len(Result[i].error) > 0:
                    ccc += 1
            # print(ll + "error:" + str(ccc) + " count:" + str(i11))
            ccc2 = 0
            i22 = 0
            for i in ListPair:
                i22 += 1
                if i.error is True:
                    ccc2 += 1
            # print(ll + "Pair error:" + str(ccc2) + " count:" + str(i22))
            # print("原句错误数量：" + str(len(listYuan)))
            if ErorShow is False:
                resTemp = "{:.1%}".format((ccc2 / i22))
                print(resTemp + "(%d/%d)" % (ccc2, i22))
            else:
                print("%d(%d)" % (ccc, len(listYuan)))


def TransRepair(ErorShow=False):
    path = "./TransRepair-Evaluated"
    print("TransRepair:")
    for item in software:
        print(item)
        for kkkk in ssss:
            paths = path + "/" + "TransRepair_" + kkkk + "_" + item + ".xlsx"
            workbook = xlrd.open_workbook(paths)
            # 根据sheet索引或者名称获取sheet内容

            sheet1 = workbook.sheet_by_index(1)  # sheet索引从0开始

            Result = {}
            listYuan = []
            ListPair = []
            for i in range(1, sheet1.nrows):
                result = sheet1.row_values(i)  # 获取第三行

                if result[0].split(":")[0] == "Issue":
                    result1 = sheet1.row_values(i + 2)

                    result2 = sheet1.row_values(i + 5)

                    error1 = []
                    for kk in range(1, len(result1) - 1):
                        if len(str(result1[kk])) >= 1:
                            error1.append(kk)

                            if listYuan.count(result1[0]) < 1:
                                listYuan.append(result1[0])

                    error2 = []
                    Result[result1[0]] = SentT(result1[0], "", error1)
                    for kk in range(1, len(result2) - 1):
                        if len(str(result2[kk])) >= 1:
                            # print(result2[0])
                            error2.append(kk)
                    Result[result2[0]] = SentT(result2[0], "", error2)

                    flag = False
                    if len(error1) > 0 or len(error2) > 0:
                        flag = True
                    ListPair.append(
                        Pair(
                            Result[result1[0]],
                            Result[result2[0]],
                            "",
                            "",
                            flag,
                            error1,
                            error2,
                            0,
                            0,
                        )
                    )

            ccc = 0
            i11 = 0
            for i in Result.keys():
                i11 += 1
                if len(Result[i].error) > 0:
                    ccc += 1
            # print(ll + "error:" + str(ccc) + " count:" + str(i11))
            ccc2 = 0
            i22 = 0
            for i in ListPair:
                i22 += 1
                if i.error is True:
                    ccc2 += 1
            # print(ll + "Pair error:" + str(ccc2) + " count:" + str(i22))
            # print("原句错误数量：" + str(len(listYuan)))
            if ErorShow is False:
                resTemp = "{:.1%}".format((ccc2 / i22))
                print(resTemp + "(%d/%d)" % (ccc2, i22))
            else:
                print("%d(%d)" % (ccc, len(listYuan)))


def STP(dataset, software, ErrorShow=False, ErrorType=False):
    path = "./results_romanian"
    print("STP:")
    ssss = dataset
    software = software
    result_all_dic = {}
    software_result = {}
    for soft in software:
        software_result[soft] = []
    for kkkk in ssss:
        result_software = {}
        for item in software:
            # print(item)
            paths = path + "/" + kkkk + "_" + item + ".xlsx"
            workbook = xlrd.open_workbook(paths)
            # 根据sheet索引或者名称获取sheet内容

            sheet1 = workbook.sheet_by_index(1)  # sheet索引从0开始

            # sheet1的名称，行数，列数
            count = 0
            SentAfter = {}
            Result = {}
            preSent = ""
            pre = ""
            ff = -1
            listYuan = []
            # excel处理
            for i in range(1, sheet1.nrows):
                result = sheet1.row_values(i)  # 获取第三行
                if result[0].split(":")[0] == "issue":
                    count = 0
                if count == 0:
                    pass
                elif count == 1:
                    pass
                elif count == 2:
                    sent = result[0]
                    SentAfter[sent] = []
                    preSent = sent
                    error = []
                    if len(str(result[6])) < 1:
                        for kk in range(1, 6):
                            if len(str(result[kk])) >= 1:
                                error.append(kk)
                                if listYuan.count(result[0]) < 1:
                                    listYuan.append(result[0])
                                if item.lower() == "google":
                                    if STPGoogleList.count(result[0]) < 1:
                                        STPGoogleList.append(result[0])
                                elif item.lower() == "google":
                                    if STPGoogleList.count(result[0]) < 1:
                                        STPGoogleList.append(result[0])
                                else:
                                    if STPBingList.count(result[0]) < 1:
                                        STPBingList.append("bing:" + result[0])
                    Result[sent] = SentT(sent, "", error)
                elif count == 3:
                    Result[preSent].trans = result[0]
                elif count == 4:
                    pass
                else:
                    if count % 2 == 1:
                        sent = result[0]
                        texts = [word for word in nlpEN.word_tokenize(sent)]
                        punc = string.punctuation
                        count_word = 0
                        for word in texts:
                            if word not in punc:  # type: ignore
                                count_word = count_word + 1
                        if count_word < 3:
                            ff = 1
                            continue
                        SentAfter[preSent].append(sent)
                        error = []
                        pre = sent
                        if len(str(result[6])) < 1:
                            for kk in range(1, 6):
                                if len(str(result[kk])) >= 1:
                                    error.append(kk)
                                    if item.lower() == "google":
                                        if STPGoogleList.count(result[0]) < 1:
                                            STPGoogleList.append(result[0])
                                    else:
                                        if STPBingList.count(result[0]) < 1:
                                            STPBingList.append(result[0])
                        Result[sent] = SentT(sent, "", error)
                    else:
                        if ff == 1:
                            ff = -1
                            continue
                        Result[pre].trans = result[0]
                count += 1

            if ErrorShow is True:
                # 错误统计
                ccc = 0
                i11 = 0
                for i in Result.keys():
                    i11 += 1
                    if len(Result[i].error) > 0:
                        ccc += 1
                res_error = "%d(%d)" % (ccc, len(listYuan))
                resList = software_result[item]
                resList.append(res_error)

            elif ErrorType is False:
                ListPair = []
                for iter in SentAfter.keys():
                    source = iter
                    pre = []
                    pre.append(source)
                    gen = SentAfter[source]
                    for new in gen:
                        LLL = pre[:]
                        pre.append(new)
                        for old in LLL:
                            flag1, worddistance = JustSub(old, new)
                            if flag1 is True:
                                distance = bow(Result[old].trans, Result[new].trans)
                                flag = False
                                if (
                                    len(Result[old].error) > 0
                                    or len(Result[new].error) > 0
                                ):
                                    flag = True
                                ListPair.append(
                                    Pair(
                                        old,
                                        new,
                                        Result[old].trans,
                                        Result[new].trans,
                                        flag,
                                        Result[old].error,
                                        Result[new].error,
                                        distance,
                                        worddistance,
                                    )
                                )

                Word_Dic = {}
                for word_distance_en in range(1, 2):
                    PairDict = {}
                    PairErrorDict = {}
                    # 匹配对错误统
                    for i in ListPair:
                        # if i.distance_word>word_distance_en:
                        #     continue
                        if i.distance not in PairDict.keys():
                            PairDict[i.distance] = 1
                        else:
                            PairDict[i.distance] = PairDict[i.distance] + 1
                        if i.error is True:
                            if i.distance not in PairErrorDict.keys():
                                PairErrorDict[i.distance] = 1
                            else:
                                PairErrorDict[i.distance] = (
                                    PairErrorDict[i.distance] + 1
                                )
                    # print(ll + "Pair error:" + str(ccc2) + " count:" + str(i22))
                    # print("原句错误数量：" + str(len(listYuan)))
                    # sent_result.append("%.2f (%d/%d)" % ((ccc2 / i22), ccc2, i22))
                    PairDict = sorted(PairDict.items())
                    PairDict = dict(PairDict)
                    PairErrorDict = sorted(PairErrorDict.items())
                    PairErrorDict = dict(PairErrorDict)
                    pair_dict_key = sorted(PairDict.keys())
                    pair_error_key = sorted(PairErrorDict.keys())
                    left_index = 0
                    right_index = 10
                    result_dic = {}
                    for dis in range(left_index, right_index + 1):
                        index_dict = -1
                        index_error = -1
                        for i in range(len(pair_dict_key)):
                            if dis > pair_dict_key[i]:
                                index_dict = i + 1
                        for i in range(len(pair_error_key)):
                            if dis > pair_error_key[i]:
                                index_error = i + 1
                        count_pair = 0
                        count_error = 0
                        for i in range(index_dict, len(pair_dict_key)):
                            count_pair = count_pair + PairDict[pair_dict_key[i]]
                        for i in range(index_error, len(pair_error_key)):
                            count_error = count_error + PairErrorDict[pair_error_key[i]]
                        if count_pair != 0:
                            res_precent = "{:.1%}".format((count_error / count_pair))
                            result_dic[dis] = res_precent + "(%d/%d)" % (
                                count_error,
                                count_pair,
                            )
                        else:
                            res_precent = "{:.1%}".format((count_error / count_pair))
                            result_dic[dis] = res_precent + "(%d/%d)" % (
                                count_error,
                                count_pair,
                            )
                    Word_Dic[word_distance_en] = result_dic
                    # print(Word_Dic)

                result_software[item] = Word_Dic
            else:
                errordict = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
                for i in Result.keys():
                    for iter in Result[i].error:
                        errordict[iter] = errordict[iter] + 1
                res_type = ""
                for i in range(1, 6):
                    res_type = res_type + str(errordict[i]) + "\t"
                resList = software_result[item]
                resList.append(res_type)
            result_all_dic[kkkk] = result_software
    if ErrorShow is True or ErrorType is True:
        for key in software_result.keys():
            print(key)
            list_temp = software_result[key]
            for item in list_temp:
                print(item)
    if ErrorShow is False:
        with open("stp.txt", "w", encoding="utf-8") as f:
            f.write(str(result_all_dic))  # dict to str


def Pairs():
    print("Pairs:")
    PatInv()
    RTI()
    SIT()
    TransRepair()
    STP(ssss, software)
    out()


def Error():
    print("Error:")
    PatInv(True)
    RTI(True)
    SIT(True)
    TransRepair(True)
    STPError()


def STPError():
    STP(dataset=ssss, software=software, ErrorShow=True)


def STPType():
    STP(dataset=ssss, software=software, ErrorShow=False, ErrorType=True)


def Overlap():
    RTI()
    STP(dataset=ssss, software=software, ErrorShow=True)
    PatInv()
    print("google")
    rti = len(list(set(RTIGoogleList).difference(PatInvGoogleList, STPGoogleList)))
    stp = len(list(set(STPGoogleList).difference(PatInvGoogleList, RTIGoogleList)))
    pat = len(list(set(PatInvGoogleList).difference(RTIGoogleList, STPGoogleList)))
    rti_stp = len(list(set(RTIGoogleList).intersection(STPGoogleList)))
    rti_Pat = len(list(set(RTIGoogleList).intersection(PatInvGoogleList)))
    rti_stp_pat = len(
        list(set(PatInvGoogleList).intersection(RTIGoogleList, STPGoogleList))
    )
    stp_pat = len(list(set(STPGoogleList).intersection(PatInvGoogleList)))
    result = (
        "RTI :"
        + str(rti)
        + "STP:"
        + str(stp)
        + "PatInv:"
        + str(pat)
        + "rti_stp:"
        + str(rti_stp - rti_stp_pat)
        + "rti_patinv"
        + str(rti_Pat - rti_stp_pat)
        + "stp_pat"
        + str(stp_pat - rti_stp_pat)
        + "stp_pat_rti"
        + str(rti_stp_pat)
    )
    print(result)
    print("bing")
    rti = len(list(set(RTIBingList).difference(PatInvBingList, STPBingList)))
    stp = len(list(set(STPBingList).difference(PatInvBingList, RTIBingList)))
    pat = len(list(set(PatInvBingList).difference(RTIBingList, STPBingList)))
    rti_stp = len(list(set(RTIBingList).intersection(STPBingList)))
    rti_Pat = len(list(set(RTIBingList).intersection(PatInvBingList)))
    rti_stp_pat = len(list(set(PatInvBingList).intersection(RTIBingList, STPBingList)))
    stp_pat = len(list(set(STPBingList).intersection(PatInvBingList)))
    result = (
        "RTI :"
        + str(rti)
        + "STP:"
        + str(stp)
        + "PatInv:"
        + str(pat)
        + "rti_stp:"
        + str(rti_stp - rti_stp_pat)
        + "rti_patinv"
        + str(rti_Pat - rti_stp_pat)
        + "stp_pat"
        + str(stp_pat - rti_stp_pat)
        + "stp_pat_rti"
        + str(rti_stp_pat)
    )
    print(result)


if __name__ == "__main__":
    # STP(
    #     dataset=["Opinion", "Tech"],
    #     software=["google", "bing", "deepL"],
    #     ErrorShow=True,
    #     ErrorType=True,
    # )
    # Pairs()
    res = bow(
        "The 70 marks consist of 30 marks.",
        "The 70 marks consist of 30 marks as score.",
    )
    print(res)
